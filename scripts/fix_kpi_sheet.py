"""
Corrige la feuille '19-KPIs & Objectifs RH' :
1. Ajoute le signe = manquant dans toutes les formules
2. Elargit les colonnes pour une bonne lisibilite
3. Ameliore la mise en page globale
"""

import sys, os
sys.path.insert(0, '/home/z/my-project/skills/xlsx/templates')
from base import (
    FONT_NAME, HEADER_BOLD, PRIMARY, PRIMARY_LIGHT, SECONDARY,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_0,
)

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

INPUT_FILE  = '/home/z/my-project/upload/Domaine1_Recrutement_Candidats (1).xlsx'
SHEET_NAME  = '19-KPIs & Objectifs RH'
FONT        = FONT_NAME

R_DEM   = "'1-Demandes Recrutement'"
R_CAN   = "'2-Base Candidats'"
R_ENT   = "'3-Planning Entretiens'"
R_EVAL  = "'4-Grille Evaluation'"
R_COUT  = "'10-Analyse Couts'"
R_PIPE  = "'18-Pipeline Candidatures'"
DS, DE = 5, 85

# ============================================================
# FORMULA BUILDERS (ALL return with leading =)
# ============================================================

def build_formula(kpi_name, month_num):
    m = month_num
    
    if kpi_name == 'Time-to-Hire (jours)':
        num = (f"SUMPRODUCT((MONTH({R_DEM}!$M${DS}:$M${DE})={m})"
               f"*({R_DEM}!$M${DS}:$M${DE}<>\"\")"
               f"*({R_DEM}!$M${DS}:$M${DE}-{R_DEM}!$C${DS}:$C${DE}))")
        den = (f"MAX(SUMPRODUCT((MONTH({R_DEM}!$M${DS}:$M${DE})={m})"
               f"*({R_DEM}!$M${DS}:$M${DE}<>\"\")),1)")
        return f"=IFERROR({num}/{den},0)"
    
    elif kpi_name == 'Cost-per-Hire (FCFA)':
        cost = (f"SUMPRODUCT((MONTH({R_COUT}!$K${DS}:$K${DE})={m})"
                f"*({R_COUT}!$F${DS}:$F${DE}+{R_COUT}!$G${DS}:$G${DE}"
                f"+{R_COUT}!$H${DS}:$H${DE}+{R_COUT}!$I${DS}:$I${DE}))")
        hires = (f"MAX(SUMPRODUCT((MONTH({R_CAN}!$AF${DS}:$AF${DE})={m})"
                 f"*({R_CAN}!$AF${DS}:$AF${DE}<>\"\")),1)")
        return f"=IFERROR({cost}/{hires},0)"
    
    elif kpi_name == 'Quality-of-Hire (/20)':
        num = (f"SUMPRODUCT((MONTH({R_EVAL}!$E${DS}:$E${DE})={m})"
               f"*({R_EVAL}!$L${DS}:$L${DE}<>\"\"))")
        den = (f"MAX(SUMPRODUCT((MONTH({R_EVAL}!$E${DS}:$E${DE})={m})"
               f"*({R_EVAL}!$L${DS}:$L${DE}<>\"\")),1)")
        return f"=IFERROR({num}/{den},0)"
    
    elif kpi_name == 'Taux acceptation offres (%)':
        acc = f"SUMPRODUCT((MONTH({R_PIPE}!$G${DS}:$G${DE})={m})*({R_PIPE}!$H${DS}:$H${DE}=\"Accepte\"))"
        ref = f"SUMPRODUCT((MONTH({R_PIPE}!$G${DS}:$G${DE})={m})*({R_PIPE}!$H${DS}:$H${DE}=\"Refuse\"))"
        den = f"MAX({acc}+{ref},1)"
        return f"=IFERROR({acc}/{den},0)"
    
    elif kpi_name == 'Taux de remplissage (%)':
        pourvue = f"SUMPRODUCT((MONTH({R_DEM}!$C${DS}:$C${DE})={m})*({R_DEM}!$L${DS}:$L${DE}=\"Pourvue\"))"
        total = f"MAX(SUMPRODUCT((MONTH({R_DEM}!$C${DS}:$C${DE})={m})*({R_DEM}!$B${DS}:$B${DE}<>\"\")),1)"
        return f"=IFERROR({pourvue}/{total},0)"
    
    elif kpi_name == 'Taux de recommandation (%)':
        reco = f"SUMPRODUCT((MONTH({R_EVAL}!$E${DS}:$E${DE})={m})*({R_EVAL}!$M${DS}:$M${DE}=\"Embauche recommandee\"))"
        total = f"MAX(SUMPRODUCT((MONTH({R_EVAL}!$E${DS}:$E${DE})={m})*({R_EVAL}!$B${DS}:$B${DE}<>\"\")),1)"
        return f"=IFERROR({reco}/{total},0)"
    
    elif kpi_name == 'Nb candidats / embauche':
        cand = f"SUMPRODUCT((MONTH({R_CAN}!$Y${DS}:$Y${DE})={m})*({R_CAN}!$B${DS}:$B${DE}<>\"\"))"
        hires = f"MAX(SUMPRODUCT((MONTH({R_CAN}!$AF${DS}:$AF${DE})={m})*({R_CAN}!$AF${DS}:$AF${DE}<>\"\")),1)"
        return f"=IFERROR({cand}/{hires},0)"
    
    elif kpi_name == 'Entretiens / embauche':
        ent = f"SUMPRODUCT((MONTH({R_ENT}!$E${DS}:$E${DE})={m})*({R_ENT}!$K${DS}:$K${DE}=\"Realise\"))"
        hires = f"MAX(SUMPRODUCT((MONTH({R_CAN}!$AF${DS}:$AF${DE})={m})*({R_CAN}!$AF${DS}:$AF${DE}<>\"\")),1)"
        return f"=IFERROR({ent}/{hires},0)"
    
    return '=0'


# ============================================================
# MAIN
# ============================================================

KPIS = [
    {'name': 'Time-to-Hire (jours)',       'unit': '0.0',     'lower_better': True},
    {'name': 'Cost-per-Hire (FCFA)',      'unit': '#,##0',   'lower_better': True},
    {'name': 'Quality-of-Hire (/20)',      'unit': '0.0',     'lower_better': False},
    {'name': 'Taux acceptation offres (%)', 'unit': '0.0%', 'lower_better': False},
    {'name': 'Taux de remplissage (%)',   'unit': '0.0%',   'lower_better': False},
    {'name': 'Taux de recommandation (%)', 'unit': '0.0%',  'lower_better': False},
    {'name': 'Nb candidats / embauche',   'unit': '0.0',     'lower_better': True},
    {'name': 'Entretiens / embauche',     'unit': '0.0',     'lower_better': True},
]

TARGETS = {
    'Time-to-Hire (jours)':       [45, 35, 30, 30],
    'Cost-per-Hire (FCFA)':      [200000, 180000, 150000, 150000],
    'Quality-of-Hire (/20)':      [13, 14, 15, 16],
    'Taux acceptation offres (%)': [0.60, 0.70, 0.75, 0.80],
    'Taux de remplissage (%)':   [0.50, 0.60, 0.70, 0.75],
    'Taux de recommandation (%)': [0.40, 0.50, 0.55, 0.60],
    'Nb candidats / embauche':   [15, 12, 10, 10],
    'Entretiens / embauche':     [5, 4, 3.5, 3],
}

MONTHS_FR = ['Jan', 'Fev', 'Mar', 'Avr', 'Mai', 'Jun', 'Jul', 'Aou', 'Sep', 'Oct', 'Nov', 'Dec']
QUARTERS = ['T1 (Jan-Mar)', 'T2 (Avr-Jun)', 'T3 (Jul-Sep)', 'T4 (Oct-Dec)']

wb = load_workbook(INPUT_FILE)
if SHEET_NAME in wb.sheetnames:
    del wb[SHEET_NAME]

ws = wb.create_sheet(SHEET_NAME)
ws.sheet_view.showGridLines = False

# ---- Column widths ----
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 28
for ci in range(3, 15):  # C to N = 12 months
    ws.column_dimensions[get_column_letter(ci)].width = 11
ws.column_dimensions['O'].width = 14  # Moy. Annuelle
for ci in range(3, 19):  # C to R for quarterly
    ws.column_dimensions[get_column_letter(ci)].width = 13

# ---- Row 1: margin ----
ws.row_dimensions[1].height = 15

# ---- Row 2: Title ----
ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=15)
c = ws.cell(row=2, column=2, value='KPIs & OBJECTIFS RH')
c.font = Font(name=FONT, size=16, bold=False, color=PRIMARY)
c.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[2].height = 32

# ---- Row 3: Subtitle ----
c = ws.cell(row=3, column=2, value='Suivi des indicateurs cles de performance — Donnees automatiques')
c.font = Font(name=FONT, size=10, bold=True, color=ACCENT_POSITIVE)

# ============================================================
# SECTION 1: KPIs MENSUELS
# ============================================================

SEC1_TITLE = 5
ws.merge_cells(start_row=SEC1_TITLE, start_column=2, end_row=SEC1_TITLE, end_column=15)
c = ws.cell(row=SEC1_TITLE, column=2, value='KPIs MENSUELS')
c.font = Font(name=FONT, size=12, bold=False, color=PRIMARY)
c.fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
for ci in range(2, 16):
    ws.cell(row=SEC1_TITLE, column=ci).fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
ws.row_dimensions[SEC1_TITLE].height = 26

# Headers
HDR_ROW = 6
headers = ['KPI'] + MONTHS_FR + ['Moy. Annuelle']
for i, h in enumerate(headers):
    col = 2 + i
    cell = ws.cell(row=HDR_ROW, column=col, value=h)
    cell.font = Font(name=FONT, size=11, bold=False, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=PRIMARY)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(bottom=Side(style='thin', color=NEUTRAL_200))
ws.row_dimensions[HDR_ROW].height = 28

# Data rows
DATA_START_ROW = 7
for kpi_idx, kpi in enumerate(KPIS):
    row = DATA_START_ROW + kpi_idx
    
    # KPI name
    cell = ws.cell(row=row, column=2, value=kpi['name'])
    cell.font = Font(name=FONT, size=11, bold=False, color=PRIMARY)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Monthly formulas (with leading =)
    for m in range(1, 13):
        col = 2 + m  # C=Jan, D=Fev, ..., N=Dec
        formula = build_formula(kpi['name'], m)
        cell = ws.cell(row=row, column=col, value=formula)
        cell.font = Font(name=FONT, size=11, bold=False, color=NEUTRAL_900)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.number_format = kpi['unit']
    
    # Annual average
    first_col = get_column_letter(3)
    last_col = get_column_letter(14)
    cell = ws.cell(row=row, column=15,
        value=f'=IFERROR(AVERAGE({first_col}{row}:{last_col}{row}),0)')
    cell.font = Font(name=FONT, size=11, bold=False, color=PRIMARY)
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
    cell.number_format = kpi['unit']
    
    ws.row_dimensions[row].height = 22

# ============================================================
# SECTION 2: OBJECTIFS TRIMESTRIELS & VARIANCE RAG
# ============================================================

SEC2_TITLE = DATA_START_ROW + len(KPIS) + 2  # row 17
ws.merge_cells(start_row=SEC2_TITLE, start_column=2, end_row=SEC2_TITLE, end_column=18)
c = ws.cell(row=SEC2_TITLE, column=2, value='OBJECTIFS TRIMESTRIELS & VARIANCE RAG')
c.font = Font(name=FONT, size=12, bold=False, color=PRIMARY)
c.fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
for ci in range(2, 19):
    ws.cell(row=SEC2_TITLE, column=ci).fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
ws.row_dimensions[SEC2_TITLE].height = 26

# Quarter header row (merged)
Q_HDR_ROW = SEC2_TITLE + 1
for qi, ql in enumerate(QUARTERS):
    start_c = 3 + qi * 4
    end_c = start_c + 3
    ws.merge_cells(start_row=Q_HDR_ROW, start_column=start_c, end_row=Q_HDR_ROW, end_column=end_c)
    cell = ws.cell(row=Q_HDR_ROW, column=start_c, value=ql)
    cell.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for cc in range(start_c, end_c + 1):
        ws.cell(row=Q_HDR_ROW, column=cc).fill = PatternFill('solid', fgColor=PRIMARY)
        ws.cell(row=Q_HDR_ROW, column=cc).font = Font(name=FONT, size=10, bold=True, color='FFFFFF')

# KPI label in quarter header
kpi_hdr = ws.cell(row=Q_HDR_ROW, column=2)
kpi_hdr.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
kpi_hdr.fill = PatternFill('solid', fgColor=PRIMARY)
ws.row_dimensions[Q_HDR_ROW].height = 24

# Sub-headers
Q_COL_HDR = Q_HDR_ROW + 1
sub_headers = ['KPI']
for _ in QUARTERS:
    sub_headers.extend(['Objectif', 'Realise', 'Ecart (%)', 'RAG'])
for i, h in enumerate(sub_headers):
    col = 2 + i
    cell = ws.cell(row=Q_COL_HDR, column=col, value=h)
    cell.font = Font(name=FONT, size=10, bold=False, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=PRIMARY)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(bottom=Side(style='thin', color=NEUTRAL_200))
ws.row_dimensions[Q_COL_HDR].height = 28

# Data rows
Q_DATA_START = Q_COL_HDR + 1

for kpi_idx, kpi in enumerate(KPIS):
    row = Q_DATA_START + kpi_idx
    monthly_row = DATA_START_ROW + kpi_idx
    targets = TARGETS[kpi['name']]
    
    # KPI name
    cell = ws.cell(row=row, column=2, value=kpi['name'])
    cell.font = Font(name=FONT, size=11, bold=False, color=PRIMARY)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    for qi in range(4):
        base_col = 3 + qi * 4
        
        # --- Objectif (blue = manual) ---
        target_cell = ws.cell(row=row, column=base_col, value=targets[qi])
        target_cell.font = Font(name=FONT, size=11, bold=False, color='0000FF')
        target_cell.alignment = Alignment(horizontal='right', vertical='center')
        if '%' in kpi['unit']:
            target_cell.number_format = '0%'
        elif kpi['unit'] == '#,##0':
            target_cell.number_format = '#,##0'
        else:
            target_cell.number_format = '0.0'
        
        # --- Realise (quarterly avg from monthly) ---
        m_start = qi * 3 + 1
        m_end = m_start + 2
        month_cells = []
        for mm in range(m_start, m_end + 1):
            mcol = get_column_letter(2 + mm)
            month_cells.append(f'{mcol}{monthly_row}')
        
        realised_formula = f"=IFERROR(AVERAGE({','.join(month_cells)}),0)"
        cell = ws.cell(row=row, column=base_col + 1, value=realised_formula)
        cell.font = Font(name=FONT, size=11, bold=False, color=NEUTRAL_900)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.number_format = kpi['unit']
        
        # --- Ecart (%) ---
        obj_ref = f"{get_column_letter(base_col)}{row}"
        real_ref = f"{get_column_letter(base_col+1)}{row}"
        
        if kpi['lower_better']:
            ecart = f"=IFERROR(({obj_ref}-{real_ref})/MAX(ABS({obj_ref}),0.001),0)"
        else:
            ecart = f"=IFERROR(({real_ref}-{obj_ref})/MAX(ABS({obj_ref}),0.001),0)"
        
        cell = ws.cell(row=row, column=base_col + 2, value=ecart)
        cell.font = Font(name=FONT, size=11, bold=False, color=NEUTRAL_900)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.number_format = '+0.0%;-0.0%;0.0%'
        
        # --- RAG ---
        ecart_ref = f"{get_column_letter(base_col+2)}{row}"
        rag_formula = f'=IF({ecart_ref}>=0.1,"OK",IF({ecart_ref}>=-0.1,"ATT","NOK"))'
        cell = ws.cell(row=row, column=base_col + 3, value=rag_formula)
        cell.font = Font(name=FONT, size=11, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[row].height = 22

# ============================================================
# CONDITIONAL FORMATTING: RAG
# ============================================================

green_fill = PatternFill('solid', fgColor='E8F5E9')
green_font = Font(name=FONT, size=11, bold=True, color=ACCENT_POSITIVE)
amber_fill = PatternFill('solid', fgColor='FEF9E7')
amber_font = Font(name=FONT, size=11, bold=True, color=ACCENT_WARNING)
red_fill = PatternFill('solid', fgColor='FDEDEC')
red_font = Font(name=FONT, size=11, bold=True, color=ACCENT_NEGATIVE)

for kpi_idx in range(len(KPIS)):
    row = Q_DATA_START + kpi_idx
    for qi in range(4):
        # Ecart column
        ecart_col = 3 + qi * 4 + 2
        ecart_letter = get_column_letter(ecart_col)
        ecart_range = f'{ecart_letter}{row}'
        
        ws.conditional_formatting.add(ecart_range,
            CellIsRule(operator='greaterThanOrEqual', formula=['0.1'],
                      fill=green_fill, font=green_font))
        ws.conditional_formatting.add(ecart_range,
            CellIsRule(operator='between', formula=['-0.1', '0.0999'],
                      fill=amber_fill, font=amber_font))
        ws.conditional_formatting.add(ecart_range,
            CellIsRule(operator='lessThan', formula=['-0.1'],
                      fill=red_fill, font=red_font))
        
        # RAG column
        rag_col = 3 + qi * 4 + 3
        rag_letter = get_column_letter(rag_col)
        rag_range = f'{rag_letter}{row}'
        
        ws.conditional_formatting.add(rag_range,
            FormulaRule(formula=[f'{ecart_letter}{row}>=0.1'],
                       fill=green_fill, font=green_font))
        ws.conditional_formatting.add(rag_range,
            FormulaRule(formula=[f'AND({ecart_letter}{row}>=-0.1,{ecart_letter}{row}<0.1)'],
                       fill=amber_fill, font=amber_font))
        ws.conditional_formatting.add(rag_range,
            FormulaRule(formula=[f'{ecart_letter}{row}<-0.1'],
                       fill=red_fill, font=red_font))

# ============================================================
# SECTION 3: LEGENDE
# ============================================================

LEG_ROW = Q_DATA_START + len(KPIS) + 2
ws.merge_cells(start_row=LEG_ROW, start_column=2, end_row=LEG_ROW, end_column=10)
c = ws.cell(row=LEG_ROW, column=2, value='LEGENDE & METHODOLOGIE')
c.font = Font(name=FONT, size=12, bold=False, color=PRIMARY)
c.fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
for ci in range(2, 11):
    ws.cell(row=LEG_ROW, column=ci).fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)

legend_items = [
    ('OK (Vert)',      'Objectif depasse de plus de 10%', ACCENT_POSITIVE),
    ('ATT (Amber)',    'A +/- 10% de l\'objectif', ACCENT_WARNING),
    ('NOK (Rouge)',    'En dessous de plus de 10%', ACCENT_NEGATIVE),
    ('', None, None),
    ('Time-to-Hire',      'Moy. (Date Pourvue - Date Demande) par mois, feuille 1', NEUTRAL_600),
    ('Cost-per-Hire',     'Somme couts / Nb embauches, feuilles 10 & 2', NEUTRAL_600),
    ('Quality-of-Hire',   'Score moyen evaluations (/25), feuille 4', NEUTRAL_600),
    ('Taux acceptation',  'Pipeline Accepte / (Accepte + Refuse), feuille 18', NEUTRAL_600),
    ('Taux remplissage',  'Demandes Pourvue / Total, feuille 1', NEUTRAL_600),
    ('Taux recommandation', 'Embauche recommandee / Total evalues, feuille 4', NEUTRAL_600),
    ('Candidats / embauche', 'Candidatures / Embauches def., feuille 2', NEUTRAL_600),
    ('Entretiens / embauche', 'Entretiens Realises / Embauches, feuilles 3 & 2', NEUTRAL_600),
    ('', None, None),
    ('Objectifs (bleu)',   'Valeurs modifiables — ajustez selon votre strategie', '0000FF'),
    ('Realise (noir)',    'Formules automatiques — ne pas modifier', NEUTRAL_900),
]

for i, (label, desc, color) in enumerate(legend_items):
    r = LEG_ROW + 1 + i
    if not label and not desc:
        continue
    cell_l = ws.cell(row=r, column=2, value=label)
    cell_l.font = Font(name=FONT, size=10, bold=True, color=color if color else NEUTRAL_900)
    
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
    cell_d = ws.cell(row=r, column=3, value=desc)
    cell_d.font = Font(name=FONT, size=10, bold=False, color=NEUTRAL_600)
    cell_d.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 20

# ============================================================
# FREEZE PANES & PRINT
# ============================================================

ws.freeze_panes = 'C7'
ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

# ============================================================
# SAVE
# ============================================================

wb.properties.creator = 'Z.ai'
wb.save(INPUT_FILE)

# Verify formulas are recognized
wb2 = load_workbook(INPUT_FILE)
ws2 = wb2['19-KPIs & Objectifs RH']
print(f"C7 data_type: {ws2['C7'].data_type} (should be 'f' for formula)")
print(f"C7 value[:40]: {str(ws2['C7'].value)[:40]}")
print(f"O7 data_type: {ws2['O7'].data_type}")
print(f"D20 data_type: {ws2['D20'].data_type} (quarterly realised)")
print(f"E20 data_type: {ws2['E20'].data_type} (ecart)")
print(f"F20 data_type: {ws2['F20'].data_type} (RAG)")
wb2.close()

print(f"\nFeuille corrigee et sauvegardee!")
print(f"Fichier: {INPUT_FILE}")
