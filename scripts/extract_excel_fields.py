import openpyxl
import json

wb = openpyxl.load_workbook('/home/z/my-project/upload/Domaine1_Recrutement_Candidats (1).xlsx', data_only=True)

result = {}
total_fields = 0
skip_sheets = ['_Lists']

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if sheet_name in skip_sheets:
        continue
    # Search rows 1-10 for header row (first row with 3+ non-empty string values)
    for row_idx in range(1, 11):
        row_vals = [cell.value for cell in ws[row_idx]]
        non_empty = [v for v in row_vals if v is not None]
        # Check if this looks like a header (mostly strings)
        str_vals = [str(v).strip() for v in non_empty if isinstance(v, str)]
        if len(str_vals) >= 3 and len(str_vals) >= len(non_empty) * 0.7:
            cleaned = [str(v).strip() for v in row_vals if v is not None]
            result[sheet_name] = cleaned
            total_fields += len(cleaned)
            print(f'{sheet_name} (row {row_idx}): {len(cleaned)} fields')
            for h in cleaned:
                print(f'  - {h}')
            break
    else:
        result[sheet_name] = []
        print(f'{sheet_name}: NO HEADERS FOUND')

print(f'\nTOTAL DATA FIELDS: {total_fields} across {len(result)} data sheets')
print(f'Nomenclatures (_Lists): 38 fields')
print(f'GRAND TOTAL: {total_fields + 38} fields')

with open('/home/z/my-project/scripts/excel_fields.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

wb.close()
