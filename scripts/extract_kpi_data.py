import openpyxl
import json
from datetime import datetime
from collections import defaultdict

def to_num(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return v
    try: return float(str(v).replace(',', '').replace(' ', ''))
    except: return 0

def to_int(v):
    return int(to_num(v))

wb = openpyxl.load_workbook('/home/z/my-project/upload/Domaine1_Recrutement_Candidats (1).xlsx', data_only=True)

def col_idx(ws, text, header_row=4):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v and text.lower() in str(v).lower():
            return c
    return None

def gc(ws, row, col):
    return ws.cell(row=row, column=col).value if col else None

# ─── 1. Demandes Recrutement (header row=4) ───
ws1 = wb['1-Demandes Recrutement']
demandes = []
for r in range(5, ws1.max_row + 1):
    dd = gc(ws1, r, 3)  # col3=Date Demande
    st = gc(ws1, r, 12) # col12=Statut
    dp = gc(ws1, r, 13) # col13=Date Pourvue
    if dd and isinstance(dd, datetime):
        st_str = str(st).strip() if st else ''
        demandes.append({
            'mois': dd.month,
            'statut': st_str,
            'pourvue': st_str.lower() == 'pourvue',
            'date_demande': dd,
            'date_pourvue': dp,
        })

# ─── 2. Base Candidats (header row=5!) ───
ws2 = wb['2-Base Candidats']
candidats = []
for r in range(6, ws2.max_row + 1):
    dc = ws2.cell(row=r, column=25).value  # col25=Date Candidature
    de = ws2.cell(row=r, column=32).value  # col32=Date Embauche Definitive
    if dc and isinstance(dc, datetime):
        candidats.append({
            'mois': dc.month,
            'date_embauche': de,
            'embauche_def': de is not None and isinstance(de, datetime),
        })

# ─── 3. Planning Entretiens (header row=4) ───
ws3 = wb['3-Planning Entretiens']
entretiens = []
for r in range(5, ws3.max_row + 1):
    de = gc(ws3, r, 5)  # col5=Date Entretien
    st = gc(ws3, r, 11) # col11=Statut
    if de and isinstance(de, datetime):
        st_str = str(st).strip() if st else ''
        entretiens.append({
            'mois': de.month,
            'statut': st_str,
            'realise': st_str.lower() == 'realise',
        })

# ─── 4. Grille Evaluation (header row=4) ───
ws4 = wb['4-Grille Evaluation']
evaluations = []
for r in range(5, ws4.max_row + 1):
    de = gc(ws4, r, 5)  # col5=Date Evaluation
    # Total is formula - compute from components col7-11 (each /5)
    comp1 = to_num(gc(ws4, r, 7))
    comp2 = to_num(gc(ws4, r, 8))
    comp3 = to_num(gc(ws4, r, 9))
    comp4 = to_num(gc(ws4, r, 10))
    comp5 = to_num(gc(ws4, r, 11))
    total = comp1 + comp2 + comp3 + comp4 + comp5
    rec = gc(ws4, r, 13)  # col13=Recommandation
    if de and isinstance(de, datetime):
        rec_str = str(rec).strip() if rec else ''
        evaluations.append({
            'mois': de.month,
            'total_25': total,
            'recommandation': rec_str,
            'recommande': 'recommande' in rec_str.lower() and 'non' not in rec_str.lower() and 'refus' not in rec_str.lower(),
        })

# ─── 10. Analyse Couts (header row=4) ───
ws10 = wb['10-Analyse Couts']
couts = []
for r in range(5, ws10.max_row + 1):
    d = ws10.cell(row=r, column=11).value  # col11=Date
    if d and isinstance(d, datetime):
        total = to_num(gc(ws10, r, 6)) + to_num(gc(ws10, r, 7)) + to_num(gc(ws10, r, 8)) + to_num(gc(ws10, r, 9))
        couts.append({'mois': d.month, 'total': total})

# ─── 18. Pipeline Candidatures (header row=4) ───
ws18 = wb['18-Pipeline Candidatures']
pipeline = []
for r in range(5, ws18.max_row + 1):
    dc = gc(ws18, r, 7)  # col7=Date Candidature
    st = gc(ws18, r, 8)  # col8=Stade Actuel
    if dc and isinstance(dc, datetime):
        st_str = str(st).strip() if st else ''
        pipeline.append({
            'mois': dc.month,
            'stade': st_str,
            'accepte': st_str.lower() == 'accepte' or 'offre accept' in st_str.lower(),
            'refuse': st_str.lower() == 'refuse' or 'refuse' in st_str.lower(),
        })

print(f"Demandes: {len(demandes)}, Candidats: {len(candidats)}, Entretiens: {len(entretiens)}, Evaluations: {len(evaluations)}, Couts: {len(couts)}, Pipeline: {len(pipeline)}")

# ─── Calculs KPIs par mois ───
kpis_monthly = {}
for m in range(1, 13):
    kpis_monthly[m] = {
        'time_to_hire': None,
        'cost_per_hire': None,
        'quality_of_hire': None,
        'taux_acceptation': None,
        'taux_remplissage': None,
        'taux_recommandation': None,
        'candidats_par_embauche': None,
        'entretiens_par_embauche': None,
    }

for m in range(1, 13):
    # 1. Time-to-Hire: moy (Date Pourvue - Date Demande) for pourvue demandes
    tths = []
    for d in demandes:
        if d['mois'] == m and d['pourvue'] and d['date_pourvue'] and isinstance(d['date_pourvue'], datetime):
            delta = (d['date_pourvue'] - d['date_demande']).days
            tths.append(delta)
    if tths:
        kpis_monthly[m]['time_to_hire'] = round(sum(tths) / len(tths), 1)

    # Also add based on date_pourvue month
    tths_pourvue = []
    for d in demandes:
        if d['pourvue'] and d['date_pourvue'] and isinstance(d['date_pourvue'], datetime) and d['date_pourvue'].month == m:
            delta = (d['date_pourvue'] - d['date_demande']).days
            tths_pourvue.append(delta)
    if tths_pourvue:
        kpis_monthly[m]['time_to_hire'] = round(sum(tths_pourvue) / len(tths_pourvue), 1)

    # 2. Cost-per-Hire: total couts / nb embauches (definitives) ce mois
    total_cout_m = sum(c['total'] for c in couts if c['mois'] == m)
    nb_embauches_m = sum(1 for c in candidats if c['mois'] == m and c['embauche_def'])
    if total_cout_m > 0 and nb_embauches_m > 0:
        kpis_monthly[m]['cost_per_hire'] = round(total_cout_m / nb_embauches_m)
    elif total_cout_m > 0:
        kpis_monthly[m]['cost_per_hire'] = round(total_cout_m)

    # 3. Quality-of-Hire: score moyen evaluations (/25 -> /20)
    qoh = []
    for e in evaluations:
        if e['mois'] == m and e['total_25'] > 0:
            qoh.append(e['total_25'] / 25 * 20)
    if qoh:
        kpis_monthly[m]['quality_of_hire'] = round(sum(qoh) / len(qoh), 1)

    # 4. Taux acceptation: Accepte / (Accepte + Refuse) in pipeline
    acc = sum(1 for p in pipeline if p['mois'] == m and p['accepte'])
    ref = sum(1 for p in pipeline if p['mois'] == m and p['refuse'])
    if acc + ref > 0:
        kpis_monthly[m]['taux_acceptation'] = round((acc / (acc + ref)) * 100, 1)

    # 5. Taux remplissage: Demandes Pourvue / Total demandes
    total_dem = sum(1 for d in demandes if d['mois'] == m)
    pourvues = sum(1 for d in demandes if d['mois'] == m and d['pourvue'])
    if total_dem > 0:
        kpis_monthly[m]['taux_remplissage'] = round((pourvues / total_dem) * 100, 1)

    # 6. Taux recommandation: Embauche recommandee / Total evalues
    total_eval = sum(1 for e in evaluations if e['mois'] == m)
    recom = sum(1 for e in evaluations if e['mois'] == m and e['recommande'])
    if total_eval > 0:
        kpis_monthly[m]['taux_recommandation'] = round((recom / total_eval) * 100, 1)

    # 7. Candidats / embauche
    nb_cand_m = sum(1 for c in candidats if c['mois'] == m)
    if nb_embauches_m > 0:
        kpis_monthly[m]['candidats_par_embauche'] = round(nb_cand_m / nb_embauches_m, 1)
    elif nb_cand_m > 0:
        kpis_monthly[m]['candidats_par_embauche'] = float(nb_cand_m)

    # 8. Entretiens / embauche
    nb_ent_m = sum(1 for e in entretiens if e['mois'] == m and e['realise'])
    if nb_embauches_m > 0:
        kpis_monthly[m]['entretiens_par_embauche'] = round(nb_ent_m / nb_embauches_m, 1)
    elif nb_ent_m > 0:
        kpis_monthly[m]['entretiens_par_embauche'] = float(nb_ent_m)

print("\nKPIs par mois:")
for m in range(1, 13):
    k = kpis_monthly[m]
    if any(v is not None for v in k.values()):
        print(f"  Mois {m}: {k}")

# Output as JSON for embedding
result = {}
for m in range(1, 13):
    k = kpis_monthly[m]
    result[str(m)] = k

with open('/home/z/my-project/scripts/kpi_data.json', 'w') as f:
    json.dump(result, f, indent=2, ensure_ascii=False)
print("\nSaved to kpi_data.json")
