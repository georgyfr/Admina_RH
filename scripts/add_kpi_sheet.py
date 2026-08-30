"""
Ajoute la feuille '19-KPIs & Objectifs RH' au fichier Excel existant.
KPIs mensuels avec formules SUMPRODUCT + Objectifs trimestriels avec RAG.
"""

import sys, os
sys.path.insert(0, '/home/z/my-project/skills/xlsx/templates')
from base import (
    FONT_NAME, HEADER_BOLD, PRIMARY, PRIMARY_LIGHT, SECONDARY,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_0,
    CF_POSITIVE_FILL, CF_NEGATIVE_FILL, CF_WARNING_FILL,
    font_title, font_header, font_subheader, font_body, font_caption,
    fill_header, fill_total, fill_data_row, border_header, border_total,
    align_title, align_header, align_number, align_text, align_date,
    setup_sheet, style_header_row, style_data_row, style_total_row,
)

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURATION
# ============================================================

INPUT_FILE  = '/home/z/my-project/upload/Domaine1_Recrutement_Candidats (1).xlsx'
OUTPUT_FILE = '/home/z/my-project/upload/Domaine1_Recrutement_Candidats (1).xlsx'

SHEET_NAME = '19-KPIs & Objectifs RH'

# Data ranges (matching existing dashboard formulas)
R_DEM   = "'1-Demandes Recrutement'"       # C=Date Demande, L=Statut, M=Date Pourvue, S=Delai
R_CAN   = "'2-Base Candidats'"              # Y=Date Candidature, Z=Statut, AA=Score, AF=Date Embauche Def.
R_ENT   = "'3-Planning Entretiens'"         # E=Date Entretien, K=Statut
R_EVAL  = "'4-Grille Evaluation'"           # E=Date Evaluation, L=Total(/25), M=Recommandation
R_COUT  = "'10-Analyse Couts'"             # F=Cout Pub, G=Cout Cabinet, H=Cout Entretiens, I=Cout Form, K=Date
R_PIPE  = "'18-Pipeline Candidatures'"     # G=Date Candidature, H=Stade

DATA_START = 5
DATA_END   = 85

# ============================================================
# KPI DEFINITIONS
# ============================================================

KPIS = [
    {
        'name': 'Time-to-Hire (jours)',
        'desc': 'Delai moyen entre demande et pourvoiement',
        'unit': '0.0',
        'lower_better': True,  # lower is better
    },
    {
        'name': 'Cost-per-Hire (FCFA)',
        'desc': 'Cout moyen par recrutement',
        'unit': '#,##0',
        'lower_better': True,
    },
    {
        'name': 'Quality-of-Hire (/20)',
        'desc': 'Score moyen evaluation candidats retenus',
        'unit': '0.0',
        'lower_better': False,  # higher is better
    },
    {
        'name': 'Taux acceptation offres (%)',
        'desc': 'Offres acceptees / Offres envoyees',
        'unit': '0.0%',
        'lower_better': False,
    },
    {
        'name': 'Taux de remplissage (%)',
        'desc': 'Postes pourvus / Total demandes',
        'unit': '0.0%',
        'lower_better': False,
    },
    {
        'name': 'Taux de recommandation (%)',
        'desc': 'Candidats recommandes / Total evalues',
        'unit': '0.0%',
        'lower_better': False,
    },
    {
        'name': 'Nb candidats / embauche',
        'desc': 'Volume de candidatures par recrutement',
        'unit': '0.0',
        'lower_better': True,
    },
    {
        'name': 'Entretiens / embauche',
        'desc': 'Efficacite du processus de selection',
        'unit': '0.0',
        'lower_better': True,
    },
]

# Default quarterly targets
TARGETS = {
    'Time-to-Hire (jours)': [45, 35, 30, 30],
    'Cost-per-Hire (FCFA)': [200000, 180000, 150000, 150000],
    'Quality-of-Hire (/20)': [13, 14, 15, 16],
    'Taux acceptation offres (%)': [0.60, 0.70, 0.75, 0.80],
    'Taux de remplissage (%)': [0.50, 0.60, 0.70, 0.75],
    'Taux de recommandation (%)': [0.40, 0.50, 0.55, 0.60],
    'Nb candidats / embauche': [15, 12, 10, 10],
    'Entretiens / embauche': [5, 4, 3.5, 3],
}

MONTHS_FR = ['Jan', 'Fev', 'Mar', 'Avr', 'Mai', 'Jun', 'Jul', 'Aou', 'Sep', 'Oct', 'Nov', 'Dec']
QUARTERS = ['T1 (Jan-Mar)', 'T2 (Avr-Jun)', 'T3 (Jul-Sep)', 'T4 (Oct-Dec)']

# ============================================================
# FORMULA BUILDERS
# ============================================================

def _sumproduct_month(range_ref, month_num, value_range=None):
    """SUMPRODUCT by month. If value_range is None, counts."""
    if value_range is None:
        return f"SUMPRODUCT((MONTH({range_ref})={month_num})*({range_ref}<>\"\"))"
    return f"SUMPRODUCT((MONTH({range_ref})={month_num})*({range_range}))"


def _countif_month_cond(range_ref, month_ref, month_num, criteria):
    """Count rows where month matches AND another column equals criteria."""
    return f"SUMPRODUCT((MONTH({month_ref})={month_num})*({range_ref}=\"{criteria}\"))"


def build_formula(kpi_name, month_num):
    """Build the KPI formula for a given month (1-12)."""
    ds, de = DATA_START, DATA_END
    m = month_num
    
    if kpi_name == 'Time-to-Hire (jours)':
        # Average of (Date Pourvue - Date Demande) where Date Pourvue is in that month
        numerator = (
            f"SUMPRODUCT((MONTH({R_DEM}!$M${ds}:$M${de})={m})"
            f"*({R_DEM}!$M${ds}:$M${de}<>\"\")"
            f"*({R_DEM}!$M${ds}:$M${de}-{R_DEM}!$C${ds}:$C${de}))"
        )
        denominator = (
            f"MAX(SUMPRODUCT((MONTH({R_DEM}!$M${ds}:$M${de})={m})"
            f"*({R_DEM}!$M${ds}:$M${de}<>\"\")),1)"
        )
        return f"IFERROR({numerator}/{denominator},0)"
    
    elif kpi_name == 'Cost-per-Hire (FCFA)':
        # Sum of all cost components / count of hires (Date Embauche Def.) in that month
        cost_sum = (
            f"SUMPRODUCT((MONTH({R_COUT}!$K${ds}:$K${de})={m})"
            f"*({R_COUT}!$F${ds}:$F${de}+{R_COUT}!$G${ds}:$G${de}"
            f"+{R_COUT}!$H${ds}:$H${de}+{R_COUT}!$I${ds}:$I${de}))"
        )
        hire_count = (
            f"MAX(SUMPRODUCT((MONTH({R_CAN}!$AF${ds}:$AF${de})={m})"
            f"*({R_CAN}!$AF${ds}:$AF${de}<>\"\")),1)"
        )
        return f"IFERROR({cost_sum}/{hire_count},0)"
    
    elif kpi_name == 'Quality-of-Hire (/20)':
        # Average evaluation total score for evaluations in that month
        numerator = (
            f"SUMPRODUCT((MONTH({R_EVAL}!$E${ds}:$E${de})={m})"
            f"*({R_EVAL}!$L${ds}:$L${de}<>\"\"))"
        )
        denominator = (
            f"MAX(SUMPRODUCT((MONTH({R_EVAL}!$E${ds}:$E${de})={m})"
            f"*({R_EVAL}!$L${ds}:$L${de}<>\"\")),1)"
        )
        return f"IFERROR({numerator}/{denominator},0)"
    
    elif kpi_name == 'Taux acceptation offres (%)':
        # Pipeline: Accepte / (Accepte + Refuse) in that month
        accepte = f"SUMPRODUCT((MONTH({R_PIPE}!$G${ds}:$G${de})={m})*({R_PIPE}!$H${ds}:$H${de}=\"Accepte\"))"
        refuse = f"SUMPRODUCT((MONTH({R_PIPE}!$G${ds}:$G${de})={m})*({R_PIPE}!$H${ds}:$H${de}=\"Refuse\"))"
        denom = f"MAX({accepte}+{refuse},1)"
        return f"IFERROR({accepte}/{denom},0)"
    
    elif kpi_name == 'Taux de remplissage (%)':
        # Demandes Pourvue / Total demandes in that month
        pourvue = f"SUMPRODUCT((MONTH({R_DEM}!$C${ds}:$C${de})={m})*({R_DEM}!$L${ds}:$L${de}=\"Pourvue\"))"
        total = f"MAX(SUMPRODUCT((MONTH({R_DEM}!$C${ds}:$C${de})={m})*({R_DEM}!$B${ds}:$B${de}<>\"\")),1)"
        return f"IFERROR({pourvue}/{total},0)"
    
    elif kpi_name == 'Taux de recommandation (%)':
        # Grille: "Embauche recommandee" / Total evaluated in that month
        reco = f"SUMPRODUCT((MONTH({R_EVAL}!$E${ds}:$E${de})={m})*({R_EVAL}!$M${ds}:$M${de}=\"Embauche recommandee\"))"
        total = f"MAX(SUMPRODUCT((MONTH({R_EVAL}!$E${ds}:$E${de})={m})*({R_EVAL}!$B${ds}:$B${de}<>\"\")),1)"
        return f"IFERROR({reco}/{total},0)"
    
    elif kpi_name == 'Nb candidats / embauche':
        # New candidatures / hires in that month
        candidats = f"SUMPRODUCT((MONTH({R_CAN}!$Y${ds}:$Y${de})={m})*({R_CAN}!$B${ds}:$B${de}<>\"\"))"
        hires = f"MAX(SUMPRODUCT((MONTH({R_CAN}!$AF${ds}:$AF${de})={m})*({R_CAN}!$AF${ds}:$AF${de}<>\"\")),1)"
        return f"IFERROR({candidats}/{hires},0)"
    
    elif kpi_name == 'Entretiens / embauche':
        # Entretiens realises / hires in that month
        entretiens = f"SUMPRODUCT((MONTH({R_ENT}!$E${ds}:$E${de})={m})*({R_ENT}!$K${ds}:$K${de}=\"Realise\"))"
        hires = f"MAX(SUMPRODUCT((MONTH({R_CAN}!$AF${ds}:$AF${de})={m})*({R_CAN}!$AF${ds}:$AF${de}<>\"\")),1)"
        return f"IFERROR({entretiens}/{hires},0)"
    
    return '0'


def build_quarterly_realised(kpi_name, start_month, end_month):
    """Build formula for quarterly realised value (average of monthly values)."""
    monthly_refs = []
    for m in range(start_month, end_month + 1):
        col = get_column_letter(3 + m - 1)  # C=Jan, D=Fev, ...
        monthly_refs.append(f'{SHEET_NAME}!{col}{{row}}')
    refs = ','.join(monthly_refs)
    return f"IFERROR(AVERAGE({refs}),0)"


# ============================================================
# MAIN SCRIPT
# ============================================================

wb = load_workbook(INPUT_FILE)

# Check if sheet already exists
if SHEET_NAME in wb.sheetnames:
    del wb[SHEET_NAME]

ws = wb.create_sheet(SHEET_NAME)

# ---- Match existing file style ----
# The existing file uses Noto Sans CJK SC (heavy font => HEADER_BOLD=False)
# with PRIMARY=1B2A4A, no alternating rows (all white fill on data)

FONT = FONT_NAME  # Noto Sans CJK SC on this system

# ---- Layout constants ----
A = 1  # margin column
B = 2  # KPI name
C = 3  # Jan / T1 fields start
# Monthly columns: C(3) to N(14) = Jan..Dec
# Quarterly section starts after monthly

MONTHLY_HEADER_ROW = 6
MONTHLY_DATA_START = 7

# ---- Row 1: margin ----
ws.row_dimensions[1].height = 15
ws.column_dimensions['A'].width = 3

# ---- Row 2: Title ----
ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=14)
title_cell = ws.cell(row=2, column=2, value='KPIs & OBJECTIFS RH')
title_cell.font = Font(name=FONT, size=16, bold=False, color=PRIMARY)
title_cell.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[2].height = 32

# ---- Row 3: Subtitle ----
subtitle_cell = ws.cell(row=3, column=2, value='Suivi des indicateurs cles de performance — Donnees automatiques')
subtitle_cell.font = Font(name=FONT, size=10, bold=True, color=ACCENT_POSITIVE)
ws.row_dimensions[3].height = 20

# ============================================================
# SECTION 1: KPIs MENSUELS
# ============================================================

SECTION1_TITLE_ROW = 5

# Section title
sec1_cell = ws.cell(row=SECTION1_TITLE_ROW, column=2, value='KPIs MENSUELS')
sec1_cell.font = Font(name=FONT, size=12, bold=False, color=PRIMARY)
sec1_cell.fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
for c in range(2, 15):
    cell = ws.cell(row=SECTION1_TITLE_ROW, column=c)
    cell.fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
ws.row_dimensions[SECTION1_TITLE_ROW].height = 26

# Header row
headers_monthly = ['KPI'] + MONTHS_FR + ['Moy. Annuelle']
for i, h in enumerate(headers_monthly):
    col = 2 + i
    cell = ws.cell(row=MONTHLY_HEADER_ROW, column=col, value=h)
    cell.font = Font(name=FONT, size=11, bold=False, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=PRIMARY)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(bottom=Side(style='thin', color=NEUTRAL_200))
ws.row_dimensions[MONTHLY_HEADER_ROW].height = 28

# KPI description row (row below header, merged with KPI name)
DESC_ROW = MONTHLY_HEADER_ROW  # We'll put desc in the KPI column

# Data rows for each KPI
for kpi_idx, kpi in enumerate(KPIS):
    row = MONTHLY_DATA_START + kpi_idx
    row_fill = PatternFill('solid', fgColor=NEUTRAL_0)
    
    # KPI name cell
    name_cell = ws.cell(row=row, column=2, value=kpi['name'])
    name_cell.font = Font(name=FONT, size=11, bold=False, color=PRIMARY)
    name_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    name_cell.fill = row_fill
    
    # Description in column B comment
    from openpyxl.comments import Comment
    name_cell.comment = Comment(kpi['desc'], 'System')
    
    # Monthly formulas
    for m in range(1, 13):
        col = 2 + m  # C=Jan(1), D=Fev(2), ..., N=Dec(12)
        formula = build_formula(kpi['name'], m)
        cell = ws.cell(row=row, column=col, value=formula)
        cell.font = Font(name=FONT, size=11, bold=False, color=NEUTRAL_900)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.fill = row_fill
        cell.number_format = kpi['unit']
    
    # Annual average (column O = 15)
    first_col = get_column_letter(3)  # C
    last_col = get_column_letter(14)  # N
    avg_cell = ws.cell(row=row, column=15,
        value=f"IFERROR(AVERAGE({first_col}{row}:{last_col}{row}),0)")
    avg_cell.font = Font(name=FONT, size=11, bold=False, color=PRIMARY)
    avg_cell.alignment = Alignment(horizontal='right', vertical='center')
    avg_cell.fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
    avg_cell.number_format = kpi['unit']
    
    ws.row_dimensions[row].height = 22

# ============================================================
# SECTION 2: OBJECTIFS TRIMESTRIELS & VARIANCE RAG
# ============================================================

SEC2_GAP = 2
SECTION2_TITLE_ROW = MONTHLY_DATA_START + len(KPIS) + SEC2_GAP

# Section title
sec2_cell = ws.cell(row=SECTION2_TITLE_ROW, column=2, value='OBJECTIFS TRIMESTRIELS & VARIANCE RAG')
sec2_cell.font = Font(name=FONT, size=12, bold=False, color=PRIMARY)
sec2_cell.fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
for c in range(2, 19):
    cell = ws.cell(row=SECTION2_TITLE_ROW, column=c)
    cell.fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
ws.row_dimensions[SECTION2_TITLE_ROW].height = 26

# Sub-header row: Quarter labels (merged)
Q_HEADER_ROW = SECTION2_TITLE_ROW + 1
q_header_labels = ['KPI']
for q_label in QUARTERS:
    q_header_labels.extend([q_label, '', '', ''])

# Simpler: just put quarter labels merged across 4 cols each
q_col = 3
for qi, ql in enumerate(QUARTERS):
    start_c = q_col + qi * 4
    end_c = start_c + 3
    ws.merge_cells(start_row=Q_HEADER_ROW, start_column=start_c, end_row=Q_HEADER_ROW, end_column=end_c)
    cell = ws.cell(row=Q_HEADER_ROW, column=start_c, value=ql)
    cell.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=PRIMARY)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for cc in range(start_c, end_c + 1):
        ws.cell(row=Q_HEADER_ROW, column=cc).fill = PatternFill('solid', fgColor=PRIMARY)
        ws.cell(row=Q_HEADER_ROW, column=cc).font = Font(name=FONT, size=10, bold=True, color='FFFFFF')

# KPI label in quarter header row
kpi_hdr = ws.cell(row=Q_HEADER_ROW, column=2)
kpi_hdr.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
kpi_hdr.fill = PatternFill('solid', fgColor=PRIMARY)

ws.row_dimensions[Q_HEADER_ROW].height = 24

# Column sub-headers: Objectif, Realise, Ecart(%), RAG
Q_COL_HEADER_ROW = Q_HEADER_ROW + 1
sub_headers = ['KPI']
for _ in QUARTERS:
    sub_headers.extend(['Objectif', 'Realise', 'Ecart (%)', 'RAG'])

for i, h in enumerate(sub_headers):
    col = 2 + i
    cell = ws.cell(row=Q_COL_HEADER_ROW, column=col, value=h)
    cell.font = Font(name=FONT, size=10, bold=False, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=PRIMARY)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(bottom=Side(style='thin', color=NEUTRAL_200))
ws.row_dimensions[Q_COL_HEADER_ROW].height = 28

# Data rows for quarterly section
Q_DATA_START = Q_COL_HEADER_ROW + 1

for kpi_idx, kpi in enumerate(KPIS):
    row = Q_DATA_START + kpi_idx
    row_fill = PatternFill('solid', fgColor=NEUTRAL_0)
    
    # KPI name
    name_cell = ws.cell(row=row, column=2, value=kpi['name'])
    name_cell.font = Font(name=FONT, size=11, bold=False, color=PRIMARY)
    name_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    name_cell.fill = row_fill
    
    monthly_row = MONTHLY_DATA_START + kpi_idx
    targets = TARGETS[kpi['name']]
    
    for qi in range(4):
        base_col = 3 + qi * 4  # 3, 7, 11, 15
        
        # --- Objectif (target) ---
        target_cell = ws.cell(row=row, column=base_col, value=targets[qi])
        target_cell.font = Font(name=FONT, size=11, bold=False, color='0000FF')  # Blue = manual input
        target_cell.alignment = Alignment(horizontal='right', vertical='center')
        target_cell.fill = row_fill
        if '%' in kpi['unit']:
            target_cell.number_format = '0%'
        elif kpi['unit'] == '#,##0':
            target_cell.number_format = '#,##0'
        else:
            target_cell.number_format = '0.0'
        
        # --- Realise (quarterly average) ---
        m_start = qi * 3 + 1  # 1, 4, 7, 10
        m_end = m_start + 2    # 3, 6, 9, 12
        
        # Build references to monthly cells
        month_cells = []
        for mm in range(m_start, m_end + 1):
            mcol = get_column_letter(2 + mm)  # C=Jan, D=Fev...
            month_cells.append(f"{SHEET_NAME}!{mcol}{monthly_row}")
        
        realised_formula = f"IFERROR(AVERAGE({','.join(month_cells)}),0)"
        realised_cell = ws.cell(row=row, column=base_col + 1, value=realised_formula)
        realised_cell.font = Font(name=FONT, size=11, bold=False, color=NEUTRAL_900)
        realised_cell.alignment = Alignment(horizontal='right', vertical='center')
        realised_cell.fill = row_fill
        realised_cell.number_format = kpi['unit']
        
        # --- Ecart (%) ---
        obj_ref = f"{get_column_letter(base_col)}{row}"
        real_ref = f"{get_column_letter(base_col+1)}{row}"
        
        if kpi['lower_better']:
            # For lower_better: positive variance = target met (actual <= target)
            # Ecart = (Objectif - Realise) / Objectif
            ecart_formula = (
                f"IFERROR(({obj_ref}-{real_ref})/MAX(ABS({obj_ref}),0.001),0)"
            )
        else:
            # For higher_better: positive variance = target met (actual >= target)
            # Ecart = (Realise - Objectif) / Objectif
            ecart_formula = (
                f"IFERROR(({real_ref}-{obj_ref})/MAX(ABS({obj_ref}),0.001),0)"
            )
        
        ecart_cell = ws.cell(row=row, column=base_col + 2, value=ecart_formula)
        ecart_cell.font = Font(name=FONT, size=11, bold=False, color=NEUTRAL_900)
        ecart_cell.alignment = Alignment(horizontal='right', vertical='center')
        ecart_cell.fill = row_fill
        ecart_cell.number_format = '+0.0%;-0.0%;0.0%'
        
        # --- RAG (formula-driven text) ---
        ecart_ref = f"{get_column_letter(base_col+2)}{row}"
        rag_formula = (
            f'=IF({ecart_ref}>=0.1,"\u2705",'
            f'IF({ecart_ref}>=-0.1,"\u26a0\ufe0f","\u274c"))'
        )
        rag_cell = ws.cell(row=row, column=base_col + 3, value=rag_formula)
        rag_cell.font = Font(name=FONT, size=14, bold=False)
        rag_cell.alignment = Alignment(horizontal='center', vertical='center')
        rag_cell.fill = row_fill
    
    ws.row_dimensions[row].height = 22

# ============================================================
# SECTION 3: LEGENDE & METHODOLOGIE
# ============================================================

LEGEND_ROW = Q_DATA_START + len(KPIS) + 2

# Legend title
leg_title = ws.cell(row=LEGEND_ROW, column=2, value='LEGENDE & METHODOLOGIE')
leg_title.font = Font(name=FONT, size=12, bold=False, color=PRIMARY)
leg_title.fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
for c in range(2, 10):
    ws.cell(row=LEGEND_ROW, column=c).fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
ws.row_dimensions[LEGEND_ROW].height = 26

# Legend items
legend_items = [
    ('RAG Vert', 'Objectif depasse de plus de 10% (ecart positif >= +10%)'),
    ('RAG Amber', 'Dans une marge de +/- 10% de l\'objectif'),
    ('RAG Rouge', 'En dessous de l\'objectif de plus de 10% (ecart negatif < -10%)'),
    ('', ''),
    ('Time-to-Hire', 'Moyenne (Date Pourvue - Date Demande) par mois, depuis feuille 1'),
    ('Cost-per-Hire', 'Somme (Publication + Cabinet + Entretiens + Formation) / Nb embauches, depuis feuille 10 & 2'),
    ('Quality-of-Hire', 'Score moyen Total (/25) des evaluations, depuis feuille 4'),
    ('Taux acceptation', 'Pipeline "Accepte" / ("Accepte" + "Refuse") par mois, depuis feuille 18'),
    ('Taux remplissage', 'Demandes "Pourvue" / Total demandes par mois, depuis feuille 1'),
    ('Taux recommandation', 'Evaluations "Embauche recommandee" / Total evalues, depuis feuille 4'),
    ('Candidats / embauche', 'Nouvelles candidatures / Embauches definitives, depuis feuille 2'),
    ('Entretiens / embauche', 'Entretiens "Realise" / Embauches definitives, depuis feuille 3 & 2'),
    ('', ''),
    ('Objectifs (bleu)', 'Valeurs modifiables — ajustez les cibles selon votre strategie'),
    ('Realise (noir)', 'Formules automatiques — ne pas modifier'),
]

for i, (label, desc) in enumerate(legend_items):
    r = LEGEND_ROW + 1 + i
    if not label and not desc:
        continue
    
    label_cell = ws.cell(row=r, column=2, value=label)
    if 'Vert' in label:
        label_cell.font = Font(name=FONT, size=10, bold=False, color=ACCENT_POSITIVE)
    elif 'Amber' in label:
        label_cell.font = Font(name=FONT, size=10, bold=False, color=ACCENT_WARNING)
    elif 'Rouge' in label:
        label_cell.font = Font(name=FONT, size=10, bold=False, color=ACCENT_NEGATIVE)
    else:
        label_cell.font = Font(name=FONT, size=10, bold=True, color=PRIMARY)
    
    desc_cell = ws.cell(row=r, column=3, value=desc)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
    desc_cell.font = Font(name=FONT, size=10, bold=False, color=NEUTRAL_600)
    desc_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    ws.row_dimensions[r].height = 20

# ============================================================
# CONDITIONAL FORMATTING: RAG on Ecart columns
# ============================================================

# Apply RAG coloring to all Ecart columns (4 quarters x 8 KPIs = 32 cells)
from openpyxl.formatting.rule import CellIsRule

green_fill = PatternFill('solid', fgColor='E8F5E9')
green_font = Font(name=FONT, size=11, bold=False, color=ACCENT_POSITIVE)
amber_fill = PatternFill('solid', fgColor='FEF9E7')
amber_font = Font(name=FONT, size=11, bold=False, color=ACCENT_WARNING)
red_fill = PatternFill('solid', fgColor='FDEDEC')
red_font = Font(name=FONT, size=11, bold=False, color=ACCENT_NEGATIVE)

for kpi_idx in range(len(KPIS)):
    row = Q_DATA_START + kpi_idx
    for qi in range(4):
        ecart_col = 3 + qi * 4 + 2  # 5, 9, 13, 17
        col_letter = get_column_letter(ecart_col)
        cell_range = f'{col_letter}{row}'
        
        # Green: >= 10%
        ws.conditional_formatting.add(cell_range,
            CellIsRule(operator='greaterThanOrEqual', formula=['0.1'],
                      fill=green_fill, font=green_font))
        # Amber: between -10% and 10%
        ws.conditional_formatting.add(cell_range,
            CellIsRule(operator='between', formula=['-0.1', '0.0999'],
                      fill=amber_fill, font=amber_font))
        # Red: < -10%
        ws.conditional_formatting.add(cell_range,
            CellIsRule(operator='lessThan', formula=['-0.1'],
                      fill=red_fill, font=red_font))

# Also add conditional formatting on RAG emoji columns
for kpi_idx in range(len(KPIS)):
    row = Q_DATA_START + kpi_idx
    for qi in range(4):
        rag_col = 3 + qi * 4 + 3  # 6, 10, 14, 18
        col_letter = get_column_letter(rag_col)
        ecart_letter = get_column_letter(rag_col - 1)
        cell_range = f'{col_letter}{row}'
        
        # Green RAG
        ws.conditional_formatting.add(cell_range,
            FormulaRule(formula=[f'{ecart_letter}{row}>=0.1'],
                       fill=green_fill, font=Font(name=FONT, size=14, color=ACCENT_POSITIVE)))
        # Amber RAG
        ws.conditional_formatting.add(cell_range,
            FormulaRule(formula=[f'AND({ecart_letter}{row}>=-0.1,{ecart_letter}{row}<0.1)'],
                       fill=amber_fill, font=Font(name=FONT, size=14, color=ACCENT_WARNING)))
        # Red RAG
        ws.conditional_formatting.add(cell_range,
            FormulaRule(formula=[f'{ecart_letter}{row}<-0.1'],
                       fill=red_fill, font=Font(name=FONT, size=14, color=ACCENT_NEGATIVE)))

# ============================================================
# COLUMN WIDTHS
# ============================================================

ws.column_dimensions['A'].width = 3    # margin
ws.column_dimensions['B'].width = 30   # KPI name
for col_idx in range(3, 15):  # C to N (months)
    ws.column_dimensions[get_column_letter(col_idx)].width = 13
ws.column_dimensions['O'].width = 15   # Moy. Annuelle

# Quarterly section columns
for col_idx in range(3, 19):  # C to R
    ws.column_dimensions[get_column_letter(col_idx)].width = 14

# ============================================================
# FREEZE PANES (freeze KPI column + headers)
# ============================================================

ws.freeze_panes = 'C7'  # Freeze column B (KPI names) and rows 1-6 (title + headers)

# ============================================================
# PRINT SETTINGS
# ============================================================

ws.sheet_properties.pageSetUpPr = None
ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

# ============================================================
# SAVE
# ============================================================

wb.properties.creator = 'Z.ai'
wb.save(OUTPUT_FILE)
print(f"Feuille '{SHEET_NAME}' ajoutee avec succes !")
print(f"Fichier sauvegarde: {OUTPUT_FILE}")
print(f"\nStructure:")
print(f"  - 8 KPIs mensuels (Jan-Dec) avec formules SUMPRODUCT")
print(f"  - Objectifs trimestriels avec variance RAG")
print(f"  - Mise en forme conditionnelle verte/ambre/rouge")
print(f"  - Legendes et methodology")
